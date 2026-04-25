from __future__ import annotations

import json
import shutil
import subprocess
import sys
from zipfile import ZipFile
from pathlib import Path

from openpyxl import load_workbook

from path_term_kit.cli import main


REPO_ROOT = Path(__file__).resolve().parents[1]


def copy_example(tmp_path: Path) -> Path:
    source = REPO_ROOT / "examples" / "fake_subspecialty"
    project = tmp_path / "fake_subspecialty"
    shutil.copytree(source, project)
    return project / "project.yaml"


def test_fake_example_runs_end_to_end(tmp_path: Path) -> None:
    config = copy_example(tmp_path)

    assert main(["validate", str(config)]) == 0
    assert main(["scan", str(config)]) == 0
    assert main(["run", str(config)]) == 0
    assert main(["qa", str(config)]) == 0

    output_dir = config.parent / "outputs"
    assert (output_dir / "evidence.xlsx").exists()
    assert (output_dir / "questionnaire.xlsx").exists()
    assert (output_dir / "deck_outline.md").exists()
    assert (output_dir / "deck_outline.html").exists()
    assert (output_dir / "scan_log.csv").exists()

    manifest = json.loads((output_dir / "run_manifest.json").read_text(encoding="utf-8"))
    assert manifest["status"] == "pass"
    assert manifest["data_gate"]["total_rows"] == 8
    assert manifest["data_gate"]["accepted_segments"] == 7

    privacy = json.loads((output_dir / "privacy_report.json").read_text(encoding="utf-8"))
    assert privacy["status"] == "pass"


def test_questionnaire_has_fixed_choices(tmp_path: Path) -> None:
    config = copy_example(tmp_path)
    assert main(["run", str(config)]) == 0

    workbook = load_workbook(config.parent / "outputs" / "questionnaire.xlsx")
    sheet = workbook["questions"]
    headers = [cell.value for cell in sheet[1]]
    assert "director_choice" in headers
    assert "owner_confirmation" in headers
    validations = list(sheet.data_validations.dataValidation)
    formulas = {validation.formula1 for validation in validations}
    assert '"同意推荐名,建议兼容保留,建议弃用,需会议讨论"' in formulas
    assert '"确认标准名,确认兼容保留,退回会议讨论,暂缓"' in formulas


def test_target_filter_excludes_non_target_organ(tmp_path: Path) -> None:
    config = copy_example(tmp_path)
    assert main(["run", str(config)]) == 0

    workbook = load_workbook(config.parent / "outputs" / "evidence.xlsx", data_only=True)
    sheet = workbook["term_summary"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    total_index = headers.index("total_hits")
    name_index = headers.index("standard_name")
    totals = {row[name_index]: row[total_index] for row in rows[1:]}
    assert totals["低级别上皮内瘤变"] == 2


def test_privacy_scan_blocks_unredacted_output(tmp_path: Path) -> None:
    config = copy_example(tmp_path)
    assert main(["run", str(config)]) == 0
    output_dir = config.parent / "outputs"
    (output_dir / "deck_outline.md").write_text("姓名:测试对象 住院号:ABC12345", encoding="utf-8")

    assert main(["qa", str(config)]) == 1
    report = json.loads((output_dir / "privacy_report.json").read_text(encoding="utf-8"))
    assert report["status"] == "fail"
    assert report["finding_count"] >= 1


def test_init_does_not_depend_on_source_template_path(tmp_path: Path) -> None:
    project_dir = tmp_path / "initialized"
    assert main(["init", "--out", str(project_dir), "--with-example"]) == 0
    assert (project_dir / "project.yaml").exists()
    assert (project_dir / "fake_subspecialty" / "project.yaml").exists()
    assert "TODO-亚专科术语标准化项目" in (project_dir / "project.yaml").read_text(encoding="utf-8")


def test_bootstrap_project_runs_before_install_style(tmp_path: Path) -> None:
    project_dir = tmp_path / "bootstrapped"
    script = REPO_ROOT / "scripts" / "bootstrap_project.py"
    result = subprocess.run(
        [sys.executable, str(script), "--out", str(project_dir)],
        check=False,
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0
    assert (project_dir / "project.yaml").exists()
    assert (project_dir / "data").is_dir()
    assert (project_dir / "outputs").is_dir()


def test_inspect_data_lists_rows_headers_and_examples(tmp_path: Path) -> None:
    source = REPO_ROOT / "examples" / "fake_subspecialty" / "reports.csv"
    attachments = tmp_path / "attachments"
    attachments.mkdir()
    shutil.copyfile(source, attachments / "reports.csv")

    assert main(["inspect-data", str(attachments)]) == 0

    report = json.loads((attachments / "inspect_data_report.json").read_text(encoding="utf-8"))
    assert report["status"] == "pass"
    assert report["total_rows"] == 8
    assert "最终检测子公司" in report["candidate_fields"]["company_field"]
    assert "单一结果" in report["candidate_fields"]["report_text_field"]
    table = report["files"][0]["tables"][0]
    assert "单一结果" in table["headers"]
    assert table["column_examples"]["单一结果"]


def test_inspect_terms_blocks_missing_required_fields(tmp_path: Path) -> None:
    bad_terms = tmp_path / "bad_terms.csv"
    bad_terms.write_text("family_id,standard_name\nT001,演示术语\n", encoding="utf-8")

    assert main(["inspect-terms", str(bad_terms)]) == 1

    report = json.loads(
        (tmp_path / "bad_terms.csv.inspect_terms.json").read_text(encoding="utf-8")
    )
    assert report["status"] == "fail"
    assert "category" in report["missing_fields"]


def test_package_results_creates_outputs_zip(tmp_path: Path) -> None:
    config = copy_example(tmp_path)
    assert main(["run", str(config)]) == 0
    assert main(["package-results", str(config)]) == 0

    zip_path = config.parent / "outputs" / "outputs.zip"
    assert zip_path.exists()
    with ZipFile(zip_path) as archive:
        names = set(archive.namelist())
    assert {
        "evidence.xlsx",
        "questionnaire.xlsx",
        "deck_outline.md",
        "deck_outline.html",
        "run_manifest.json",
        "privacy_report.json",
        "scan_log.csv",
    }.issubset(names)


def test_inspect_data_auto_skips_term_catalog_when_mixed(tmp_path: Path) -> None:
    attachments = tmp_path / "attachments"
    attachments.mkdir()
    shutil.copyfile(REPO_ROOT / "examples" / "fake_subspecialty" / "reports.csv", attachments / "reports.csv")
    shutil.copyfile(REPO_ROOT / "examples" / "fake_subspecialty" / "terms.csv", attachments / "terms.csv")

    assert main(["inspect-data", str(attachments)]) == 0

    report = json.loads((attachments / "inspect_data_report.json").read_text(encoding="utf-8"))
    assert report["discovered_file_count"] == 2
    assert report["file_count"] == 1
    assert report["total_rows"] == 8
    assert report["skipped_files"][0]["reason"] == "looks_like_term_catalog"


def test_create_project_from_confirmed_chat_fields_runs_pipeline(tmp_path: Path) -> None:
    project = tmp_path / "created_project"
    reports = REPO_ROOT / "examples" / "fake_subspecialty" / "reports.csv"
    terms = REPO_ROOT / "examples" / "fake_subspecialty" / "terms.csv"

    assert (
        main(
            [
                "create-project",
                "--out",
                str(project),
                "--term-file",
                str(terms),
                "--report-file",
                str(reports),
                "--project-name",
                "聊天创建演示",
                "--subspecialty",
                "消化道",
                "--company-field",
                "最终检测子公司",
                "--report-text-field",
                "单一结果",
                "--context-field",
                "诊断",
                "--context-field",
                "送检材料",
                "--include-term",
                "胃",
                "--include-term",
                "胃窦",
                "--include-term",
                "胃体",
                "--include-term",
                "贲门",
                "--exclude-term",
                "结肠",
            ]
        )
        == 0
    )
    assert (project / "project.yaml").exists()
    assert main(["validate", str(project / "project.yaml")]) == 0
    assert main(["run", str(project / "project.yaml")]) == 0
    assert main(["package-results", str(project / "project.yaml")]) == 0
    assert (project / "outputs" / "outputs.zip").exists()


def test_package_results_fails_when_outputs_missing(tmp_path: Path) -> None:
    config = copy_example(tmp_path)

    assert main(["package-results", str(config)]) == 2
