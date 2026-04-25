from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from path_term_kit.aggregate import AggregationResult, FamilyEvidence, summarize_family
from path_term_kit.config import ProjectConfig, output_path


DIRECTOR_CHOICES = ("同意推荐名", "建议兼容保留", "建议弃用", "需会议讨论")
OWNER_CHOICES = ("确认标准名", "确认兼容保留", "退回会议讨论", "暂缓")


def write_evidence_workbook(config: ProjectConfig, result: AggregationResult) -> Path:
    path = output_path(config, "evidence_workbook")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "term_summary"

    summary_headers = [
        "family_id",
        "category",
        "standard_name",
        "priority",
        "total_hits",
        "nonzero_companies",
        "max_company",
        "max_company_share",
        "conflict_summary",
        "source_basis",
        "compatible_names",
        "deprecated_or_discuss",
        "decision_question",
    ]
    _append_table(summary, summary_headers, [summarize_family(item, config.companies) for item in result.evidence.values()])

    counts = workbook.create_sheet("company_counts")
    count_headers = ["family_id", "standard_name", *config.companies, "UNKNOWN", "total_hits"]
    count_rows = []
    for item in result.evidence.values():
        row = {
            "family_id": item.family.family_id,
            "standard_name": item.family.standard_name,
            "total_hits": item.total_hits,
            "UNKNOWN": item.company_counts.get("UNKNOWN", 0),
        }
        row.update({company: item.company_counts.get(company, 0) for company in config.companies})
        count_rows.append(row)
    _append_table(counts, count_headers, count_rows)

    variants = workbook.create_sheet("variant_counts")
    variant_headers = ["family_id", "standard_name", "variant", *config.companies, "UNKNOWN", "total_hits"]
    variant_rows = []
    for item in result.evidence.values():
        for variant, company_counter in item.variant_counts.items():
            row = {
                "family_id": item.family.family_id,
                "standard_name": item.family.standard_name,
                "variant": variant,
                "UNKNOWN": company_counter.get("UNKNOWN", 0),
                "total_hits": sum(company_counter.values()),
            }
            row.update({company: company_counter.get(company, 0) for company in config.companies})
            variant_rows.append(row)
    _append_table(variants, variant_headers, variant_rows)

    examples = workbook.create_sheet("examples")
    example_rows = []
    for item in result.evidence.values():
        for company, snippets in item.examples.items():
            for snippet in snippets:
                example_rows.append(
                    {
                        "family_id": item.family.family_id,
                        "standard_name": item.family.standard_name,
                        "company": company,
                        "deidentified_example": snippet,
                    }
                )
    _append_table(examples, ["family_id", "standard_name", "company", "deidentified_example"], example_rows)

    gate = workbook.create_sheet("data_gate")
    _append_table(
        gate,
        ["key", "value"],
        [
            {"key": "total_rows", "value": result.total_rows},
            {"key": "total_non_empty_rows", "value": result.total_non_empty_rows},
            {"key": "accepted_segments", "value": result.accepted_segments},
            {"key": "errors", "value": "\n".join(result.errors)},
            {"key": "warnings", "value": "\n".join(result.warnings)},
        ],
    )

    scan = workbook.create_sheet("scan_log")
    _append_table(
        scan,
        ["file", "table", "rows", "non_empty_rows", "headers", "digest", "error"],
        [log.as_dict() for log in result.logs],
    )
    _autosize(workbook)
    workbook.save(path)
    return path


def write_questionnaire_workbook(config: ProjectConfig, result: AggregationResult) -> Path:
    path = output_path(config, "questionnaire_workbook")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "questions"
    headers = [
        "family_id",
        "category",
        "recommended_standard_name",
        "source_basis",
        "compatible_names",
        "deprecated_or_discuss",
        "company_counts",
        "conflict_summary",
        "deidentified_examples",
        "director_choice",
        "director_reason",
        "owner_confirmation",
    ]
    rows = [_question_row(item, config) for item in result.evidence.values()]
    _append_table(sheet, headers, rows)
    director_validation = DataValidation(type="list", formula1=f'"{",".join(DIRECTOR_CHOICES)}"')
    owner_validation = DataValidation(type="list", formula1=f'"{",".join(OWNER_CHOICES)}"')
    sheet.add_data_validation(director_validation)
    sheet.add_data_validation(owner_validation)
    if rows:
        director_validation.add(f"J2:J{len(rows) + 1}")
        owner_validation.add(f"L2:L{len(rows) + 1}")
    _autosize(workbook)
    workbook.save(path)
    return path


def _question_row(item: FamilyEvidence, config: ProjectConfig) -> dict[str, object]:
    summary = summarize_family(item, config.companies)
    company_counts = "；".join(
        [f"{company}:{item.company_counts.get(company, 0)}" for company in config.companies]
    )
    examples = []
    for company in config.companies:
        snippets = item.examples.get(company, [])
        if snippets:
            examples.append(f"{company}: {snippets[0]}")
    return {
        "family_id": item.family.family_id,
        "category": item.family.category,
        "recommended_standard_name": item.family.standard_name,
        "source_basis": item.family.source_basis,
        "compatible_names": item.family.compatible_names,
        "deprecated_or_discuss": item.family.deprecated_or_discuss,
        "company_counts": company_counts,
        "conflict_summary": summary["conflict_summary"],
        "deidentified_examples": "\n".join(examples),
        "director_choice": "",
        "director_reason": "",
        "owner_confirmation": "",
    }


def _append_table(sheet, headers: list[str], rows: list[dict[str, object]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.style = "Headline 4"


def _autosize(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            letter = column[0].column_letter
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 60))
            sheet.column_dimensions[letter].width = max(max_length + 2, 12)

