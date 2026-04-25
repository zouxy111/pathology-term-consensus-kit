from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


REQUIRED_TERM_FIELDS = (
    "family_id",
    "category",
    "standard_name",
    "source_basis",
    "compatible_names",
    "deprecated_or_discuss",
    "patterns",
    "priority",
    "decision_question",
)


class TermCatalogError(ValueError):
    """Raised when a terminology catalog is incomplete."""


@dataclass(frozen=True)
class PatternSpec:
    label: str
    regex: re.Pattern[str]


@dataclass(frozen=True)
class TermFamily:
    family_id: str
    category: str
    standard_name: str
    source_basis: str
    compatible_names: str
    deprecated_or_discuss: str
    priority: str
    decision_question: str
    patterns: tuple[PatternSpec, ...]


def load_term_catalog(path: Path) -> list[TermFamily]:
    rows = _read_rows(path)
    if not rows:
        raise TermCatalogError("Term catalog is empty.")
    missing = [field for field in REQUIRED_TERM_FIELDS if field not in rows[0]]
    if missing:
        raise TermCatalogError("Term catalog missing fields: " + ", ".join(missing))
    families = [_row_to_family(row, index) for index, row in enumerate(rows, start=2)]
    ids = [family.family_id for family in families]
    duplicates = sorted({family_id for family_id in ids if ids.count(family_id) > 1})
    if duplicates:
        raise TermCatalogError("Duplicate family_id values: " + ", ".join(duplicates))
    return families


def _read_rows(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [{key: value or "" for key, value in row.items()} for row in csv.DictReader(handle)]
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.worksheets[0]
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value).strip() if value is not None else "" for value in next(rows, [])]
            result = []
            for values in rows:
                result.append(
                    {
                        header: str(values[index]) if index < len(values) and values[index] is not None else ""
                        for index, header in enumerate(headers)
                    }
                )
            return result
        finally:
            workbook.close()
    raise TermCatalogError(f"Unsupported term catalog file type: {path.suffix}")


def _row_to_family(row: dict[str, str], row_number: int) -> TermFamily:
    family_id = row.get("family_id", "").strip()
    standard_name = row.get("standard_name", "").strip()
    if not family_id:
        raise TermCatalogError(f"Term catalog row {row_number} missing family_id.")
    if not standard_name:
        raise TermCatalogError(f"Term catalog row {row_number} missing standard_name.")
    pattern_text = row.get("patterns", "")
    fallback_terms = [standard_name, row.get("compatible_names", ""), row.get("deprecated_or_discuss", "")]
    patterns = tuple(_compile_patterns(pattern_text, fallback_terms))
    if not patterns:
        raise TermCatalogError(f"Term catalog row {row_number} has no usable patterns.")
    return TermFamily(
        family_id=family_id,
        category=row.get("category", "").strip(),
        standard_name=standard_name,
        source_basis=row.get("source_basis", "").strip(),
        compatible_names=row.get("compatible_names", "").strip(),
        deprecated_or_discuss=row.get("deprecated_or_discuss", "").strip(),
        priority=row.get("priority", "").strip() or "main",
        decision_question=row.get("decision_question", "").strip(),
        patterns=patterns,
    )


def _compile_patterns(pattern_text: str, fallback_terms: Iterable[str]) -> list[PatternSpec]:
    parts = _split_terms(pattern_text)
    if not parts:
        for value in fallback_terms:
            parts.extend(_split_terms(value))
    specs = []
    seen = set()
    for part in parts:
        if part in seen:
            continue
        seen.add(part)
        if part.startswith("regex:"):
            pattern = part.removeprefix("regex:").strip()
            if pattern:
                specs.append(PatternSpec(label=part, regex=re.compile(pattern, flags=re.IGNORECASE)))
        else:
            specs.append(PatternSpec(label=part, regex=re.compile(re.escape(part), flags=re.IGNORECASE)))
    return specs


def _split_terms(value: str) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in re.split(r"[|｜;；,，、\n]+", value) if part.strip()]

