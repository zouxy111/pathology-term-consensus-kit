from __future__ import annotations

import csv
import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook


class ReaderError(RuntimeError):
    """Raised when an input table cannot be read safely."""


@dataclass(frozen=True)
class TableRow:
    source_file: Path
    table_name: str
    row_number: int
    values: dict[str, str]


@dataclass
class ScanLog:
    file: str
    table: str
    rows: int = 0
    non_empty_rows: int = 0
    headers: tuple[str, ...] = ()
    digest: str = ""
    error: str = ""

    def as_dict(self) -> dict[str, str | int]:
        return {
            "file": self.file,
            "table": self.table,
            "rows": self.rows,
            "non_empty_rows": self.non_empty_rows,
            "headers": "|".join(self.headers),
            "digest": self.digest,
            "error": self.error,
        }


def iter_table_rows(path: Path) -> Iterator[tuple[TableRow | None, ScanLog]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from _iter_csv(path)
    elif suffix in {".xlsx", ".xlsm"}:
        yield from _iter_xlsx(path)
    elif suffix == ".xls":
        log = ScanLog(file=str(path), table=path.name, error="Unsupported .xls; convert to .xlsx or CSV.")
        yield None, log
    else:
        log = ScanLog(file=str(path), table=path.name, error=f"Unsupported file type: {suffix}")
        yield None, log


def read_headers(path: Path) -> list[tuple[str, tuple[str, ...], str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            headers = next(reader, [])
        return [(path.name, _normalize_headers(headers), "")]
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            result = []
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=True)
                headers = next(rows, [])
                result.append((sheet.title, _normalize_headers(headers), ""))
            return result
        finally:
            workbook.close()
    return [(path.name, tuple(), f"Unsupported file type: {suffix}")]


def _iter_csv(path: Path) -> Iterator[tuple[TableRow | None, ScanLog]]:
    digest = hashlib.sha256()
    log = ScanLog(file=str(path), table=path.name)
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            log.headers = tuple(reader.fieldnames or ())
            for row_number, raw in enumerate(reader, start=2):
                values = _stringify_dict(raw)
                _update_digest(digest, values)
                log.rows += 1
                if any(value.strip() for value in values.values()):
                    log.non_empty_rows += 1
                yield TableRow(path, path.name, row_number, values), log
        log.digest = digest.hexdigest()
        yield None, log
    except Exception as exc:  # pragma: no cover - exact CSV exceptions vary by platform.
        log.error = str(exc)
        yield None, log


def _iter_xlsx(path: Path) -> Iterator[tuple[TableRow | None, ScanLog]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - exact openpyxl exceptions vary by file.
        yield None, ScanLog(file=str(path), table=path.name, error=str(exc))
        return
    try:
        for sheet in workbook.worksheets:
            digest = hashlib.sha256()
            log = ScanLog(file=str(path), table=sheet.title)
            try:
                rows = sheet.iter_rows(values_only=True)
                headers = _normalize_headers(next(rows, []))
                log.headers = headers
                for row_number, values_tuple in enumerate(rows, start=2):
                    values = {
                        header: _to_string(values_tuple[index] if index < len(values_tuple) else "")
                        for index, header in enumerate(headers)
                    }
                    _update_digest(digest, values)
                    log.rows += 1
                    if any(value.strip() for value in values.values()):
                        log.non_empty_rows += 1
                    yield TableRow(path, sheet.title, row_number, values), log
                log.digest = digest.hexdigest()
                yield None, log
            except Exception as exc:  # pragma: no cover - defensive per-sheet logging.
                log.error = str(exc)
                yield None, log
    finally:
        workbook.close()


def _normalize_headers(headers: object) -> tuple[str, ...]:
    return tuple(_to_string(value).strip() for value in headers or [])


def _stringify_dict(raw: dict[str, object]) -> dict[str, str]:
    return {_to_string(key).strip(): _to_string(value) for key, value in raw.items() if key is not None}


def _to_string(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _update_digest(digest: "hashlib._Hash", values: dict[str, str]) -> None:
    for key in sorted(values):
        digest.update(key.encode("utf-8", errors="ignore"))
        digest.update(b"\0")
        digest.update(values[key].encode("utf-8", errors="ignore"))
        digest.update(b"\0")
    digest.update(b"\n")

